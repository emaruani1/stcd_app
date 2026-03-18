"""
Upload STCD Contacts and Transactions to DynamoDB.
Run: py backend/upload_data.py
"""
import openpyxl
import boto3
import json
import uuid
import re
from decimal import Decimal
from datetime import datetime

session = boto3.Session(profile_name='stcd', region_name='us-east-2')
dynamodb = session.resource('dynamodb')

members_table = dynamodb.Table('stcd_members')
transactions_table = dynamodb.Table('stcd_transactions')
pledges_table = dynamodb.Table('stcd_pledges')
settings_table = dynamodb.Table('stcd_settings')

CONTACTS_FILE = 'C:/Users/elima/Downloads/STCD Contacts.xlsx'
TRANSACTIONS_FILE = 'C:/Users/elima/Downloads/STCD Transactions 2026.xlsx'

# ========== UPLOAD CONTACTS ==========
def upload_contacts():
    wb = openpyxl.load_workbook(CONTACTS_FILE)
    ws = wb['Contacts']

    count = 0
    name_to_id = {}  # For transaction matching later

    with members_table.batch_writer() as batch:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            acct_id, contact_type, _full, last_name, first_name, formal, dear, addr1, addr2, city, state, zip_code, country, email, phone = row

            if not last_name:
                continue

            member_id = str(int(acct_id)) if acct_id else str(uuid.uuid4())[:8]
            first_name = str(first_name).strip() if first_name else ''
            last_name = str(last_name).strip() if last_name else ''

            item = {
                'memberId': member_id,
                'lastName': last_name,
                'firstName': first_name,
                'contactType': str(contact_type).strip() if contact_type else '',
                'email': str(email).strip() if email else '',
                'phone': str(phone).strip() if phone else '',
                'address': str(addr1).strip() if addr1 else '',
                'addressLine2': str(addr2).strip() if addr2 else '',
                'city': str(city).strip() if city else '',
                'state': str(state).strip() if state else '',
                'zip': str(zip_code).strip() if zip_code else '',
                'country': str(country).strip() if country else '',
                'formalSalutation': str(formal).strip() if formal else '',
                'dearWho': str(dear).strip() if dear else '',
                'balance': Decimal('0'),
                'aliases': [],
                'yahrzeits': [],
                'children': [],
                'memberSince': '',
                'membershipType': '',
                'membershipPlan': '',
                'gender': '',
                'spouseName': '',
                'spouseGender': '',
                'spouseDob': '',
                'dob': '',
                'marriageDate': '',
            }

            # Remove empty strings from item to keep DynamoDB clean
            item = {k: v for k, v in item.items() if v != '' and v != [] or k in ('memberId', 'lastName', 'firstName', 'balance')}

            batch.put_item(Item=item)
            count += 1

            # Build name lookup for transaction matching
            # Key: "LastName FirstName" lowercase
            lookup_key = f"{last_name} {first_name}".strip().lower()
            name_to_id[lookup_key] = member_id
            # Also add just last name for partial matching
            name_to_id[last_name.lower()] = member_id

    print(f"Uploaded {count} contacts to stcd_members")
    return name_to_id


# ========== MATCH TRANSACTION ACCOUNT TO MEMBER ID ==========
def find_member_id(account_name, name_to_id):
    """Try to match a transaction account name to a member ID."""
    if not account_name:
        return None

    name = account_name.strip()
    name_lower = name.lower()

    # Direct match
    if name_lower in name_to_id:
        return name_to_id[name_lower]

    # Try without comma (e.g., "Bolurian, Daniel" -> "Bolurian Daniel")
    no_comma = name_lower.replace(',', '')
    if no_comma in name_to_id:
        return name_to_id[no_comma]

    # Try just the first word (last name)
    parts = name_lower.split()
    if parts and parts[0] in name_to_id:
        return name_to_id[parts[0]]

    # Try without the comma and just last name
    no_comma_parts = no_comma.split()
    if no_comma_parts and no_comma_parts[0] in name_to_id:
        return name_to_id[no_comma_parts[0]]

    # Skip Zelle reference strings (contain "ON xx/xx REF #")
    if 'REF #' in name or ' ON ' in name:
        # Try to extract a real name from the end
        # e.g., "TEHILA KABILIO LOPO ON 02/07 REF # PNCAA0XNX44K SHARON KABILIO"
        # Try the part after the last ref hash
        after_ref = name.split('REF #')[-1].strip() if 'REF #' in name else ''
        if after_ref:
            # Remove the ref code (first word after REF #)
            words = after_ref.split()
            if len(words) > 1:
                possible_name = ' '.join(words[1:]).lower()
                if possible_name in name_to_id:
                    return name_to_id[possible_name]
                # Try just last word as last name
                for w in words[1:]:
                    if w.lower() in name_to_id:
                        return name_to_id[w.lower()]

        # Try the name before "ON"
        before_on = name.split(' ON ')[0].strip().lower()
        parts = before_on.split()
        # Try last word as last name
        if parts:
            for p in parts:
                if p in name_to_id:
                    return name_to_id[p]

    return None


# ========== MAP PAYMENT TYPE ==========
def map_payment_type(txn_type):
    """Map spreadsheet TYPE to our app payment types."""
    mapping = {
        'MEMBERSHIP': 'membership',
        'PLEDGE': 'pledge',         # This is a pledge record, not a transaction
        'PLEDGE PAYMENT': 'pledge', # Payment toward a pledge
        'DONATION': 'donation',
        'PURCHASE': 'purchase',
    }
    return mapping.get(txn_type, 'donation')


# ========== MAP SOURCE TO PAYMENT METHOD ==========
def map_source(source):
    """Map spreadsheet SOURCE to payment method."""
    if not source:
        return ''
    mapping = {
        'FIDELITY': 'Fidelity',
        'ZELLE': 'Zelle',
        'CHECK': 'Check',
        'CHEQUE': 'Check',
    }
    return mapping.get(source.upper(), source)


# ========== UPLOAD TRANSACTIONS & PLEDGES ==========
def upload_transactions(name_to_id):
    wb = openpyxl.load_workbook(TRANSACTIONS_FILE)
    ws = wb['Transactions']

    txn_count = 0
    pledge_count = 0
    unmatched = []

    # We need to create unmatched members too
    unmatched_members_created = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        source, date, direction, account_name, txn_type, category, description, amount = row

        if not account_name or not txn_type:
            continue

        # Find or create member
        member_id = find_member_id(account_name, name_to_id)

        if not member_id:
            # Create a new member for unmatched accounts
            if account_name not in unmatched_members_created:
                new_id = str(900000 + len(unmatched_members_created))
                # Parse name - usually "LastName FirstName" format
                parts = account_name.strip().split(None, 1)
                last = parts[0] if parts else account_name
                first = parts[1] if len(parts) > 1 else ''

                # Skip Zelle reference entries - assign to generic
                if 'REF #' in account_name or ' ON ' in account_name:
                    unmatched.append(account_name)
                    continue

                members_table.put_item(Item={
                    'memberId': new_id,
                    'lastName': last,
                    'firstName': first,
                    'balance': Decimal('0'),
                })
                name_to_id[account_name.strip().lower()] = new_id
                unmatched_members_created[account_name] = new_id
                member_id = new_id
            else:
                member_id = unmatched_members_created[account_name]

        # Parse date
        if isinstance(date, datetime):
            date_str = date.strftime('%Y-%m-%d')
        else:
            date_str = str(date)[:10] if date else '2026-01-01'

        year_month = date_str[:7]  # "2026-01"
        amount_dec = Decimal(str(amount)) if amount else Decimal('0')
        payment_method = map_source(source)

        if txn_type == 'PLEDGE':
            # This is a pledge record (promise to pay), goes to pledges table
            pledge_id = f"PLG#{date_str}#{uuid.uuid4().hex[:8]}"

            # Map description to pledgeType
            pledge_type = str(description).strip() if description else ''
            occasion = str(category).strip() if category else ''

            pledges_table.put_item(Item={
                'memberId': member_id,
                'pledgeId': pledge_id,
                'date': date_str,
                'pledgeType': pledge_type,
                'occasion': occasion,
                'description': pledge_type or 'Pledge',
                'amount': amount_dec,
                'paidAmount': Decimal('0'),
                'paid': False,
                'canceled': False,
                'paymentMethod': '',
                'category': 'pledge',
            })
            pledge_count += 1
        else:
            # This is an actual transaction (payment)
            txn_id = f"TXN#{date_str}#{uuid.uuid4().hex[:8]}"
            payment_type = map_payment_type(txn_type)

            desc = str(description).strip() if description else txn_type
            if txn_type == 'MEMBERSHIP':
                desc = 'Membership Payment'
            elif txn_type == 'PLEDGE PAYMENT' and not description:
                desc = 'Pledge Payment'
            elif txn_type == 'DONATION' and not description:
                desc = 'Donation'
            elif txn_type == 'PURCHASE' and not description:
                desc = 'Purchase'

            transactions_table.put_item(Item={
                'memberId': member_id,
                'transactionId': txn_id,
                'txnDate': date_str,
                'yearMonth': year_month,
                'date': date_str,
                'source': source or '',
                'paymentType': payment_type,
                'category': str(category).strip() if category else '',
                'description': desc,
                'amount': amount_dec,
                'method': payment_method,
            })
            txn_count += 1

    print(f"Uploaded {txn_count} transactions to stcd_transactions")
    print(f"Uploaded {pledge_count} pledges to stcd_pledges")
    print(f"Created {len(unmatched_members_created)} new members for unmatched accounts")
    if unmatched:
        print(f"Skipped {len(unmatched)} Zelle reference entries (no match):")
        for u in unmatched:
            print(f"  - {u}")


# ========== UPLOAD SETTINGS ==========
def upload_settings():
    # Pledge types (extracted from transaction descriptions)
    pledge_types = [
        {'id': 'opening_ark', 'label': 'Opening the Ark', 'occasions': ['Shabbat', 'Pessah 1', 'Pessah 2', 'Sukkot 1', 'Sukkot 2', 'Rosh Hashana 1', 'Rosh Hashana 2', 'Yom Kippur', 'Simcha Torah']},
        {'id': 'taking_out_sefer', 'label': 'Taking out Sefer Torah', 'occasions': ['Shabbat', 'Pessah 1', 'Pessah 2', 'Sukkot 1', 'Sukkot 2', 'Rosh Hashana 1', 'Rosh Hashana 2', 'Yom Kippur', 'Simcha Torah']},
        {'id': 'pointing_sefer', 'label': 'Pointing at the Sefer', 'occasions': ['Shabbat', 'Pessah 1', 'Pessah 2', 'Sukkot 1', 'Sukkot 2', 'Rosh Hashana 1', 'Rosh Hashana 2', 'Yom Kippur', 'Simcha Torah']},
        {'id': 'hagbaha', 'label': 'Hagbaa', 'occasions': ['Shabbat', 'Pessah 1', 'Pessah 2', 'Sukkot 1', 'Sukkot 2', 'Rosh Hashana 1', 'Rosh Hashana 2', 'Yom Kippur', 'Simcha Torah']},
        {'id': 'aliyah_1', 'label': '1st Aliah / Kohen', 'occasions': ['Shabbat', 'Pessah 1', 'Pessah 2', 'Sukkot 1', 'Sukkot 2', 'Rosh Hashana 1', 'Rosh Hashana 2', 'Yom Kippur']},
        {'id': 'aliyah_2', 'label': '2nd Aliah / Levi', 'occasions': ['Shabbat', 'Pessah 1', 'Pessah 2', 'Sukkot 1', 'Sukkot 2', 'Rosh Hashana 1', 'Rosh Hashana 2', 'Yom Kippur']},
        {'id': 'aliyah_3', 'label': '3rd Aliah', 'occasions': ['Shabbat', 'Pessah 1', 'Pessah 2', 'Sukkot 1', 'Sukkot 2', 'Rosh Hashana 1', 'Rosh Hashana 2', 'Yom Kippur']},
        {'id': 'aliyah_4', 'label': '4th Aliah', 'occasions': ['Shabbat', 'Pessah 1', 'Pessah 2', 'Sukkot 1', 'Sukkot 2', 'Rosh Hashana 1', 'Rosh Hashana 2', 'Yom Kippur']},
        {'id': 'aliyah_5', 'label': '5th Aliah', 'occasions': ['Shabbat', 'Pessah 1', 'Pessah 2', 'Sukkot 1', 'Sukkot 2', 'Rosh Hashana 1', 'Rosh Hashana 2', 'Yom Kippur']},
        {'id': 'aliyah_6', 'label': '6th Aliah', 'occasions': ['Shabbat', 'Rosh Hashana 1', 'Rosh Hashana 2', 'Yom Kippur']},
        {'id': 'aliyah_7', 'label': '7th Aliah / Mashlim', 'occasions': ['Shabbat']},
        {'id': 'maftir', 'label': 'Maftir', 'occasions': ['Shabbat', 'Pessah 1', 'Pessah 2', 'Sukkot 1', 'Sukkot 2', 'Rosh Hashana 1', 'Rosh Hashana 2', 'Yom Kippur', 'Simcha Torah']},
        {'id': 'kiddush', 'label': 'Kiddush', 'occasions': ['Shabbat']},
        {'id': 'seuda', 'label': 'Seuda Shelishit', 'occasions': ['Shabbat']},
    ]

    occasions = [
        {'id': 'shabbat', 'label': 'Shabbat'},
        {'id': 'pessah_1', 'label': 'Pessah 1'},
        {'id': 'pessah_2', 'label': 'Pessah 2'},
        {'id': 'sukkot_1', 'label': 'Sukkot 1'},
        {'id': 'sukkot_2', 'label': 'Sukkot 2'},
        {'id': 'rosh_hashana_1', 'label': 'Rosh Hashana 1'},
        {'id': 'rosh_hashana_2', 'label': 'Rosh Hashana 2'},
        {'id': 'yom_kippur', 'label': 'Yom Kippur'},
        {'id': 'simcha_torah', 'label': 'Simcha Torah'},
    ]

    payment_methods = [
        {'id': 'fidelity', 'label': 'Fidelity'},
        {'id': 'zelle', 'label': 'Zelle'},
        {'id': 'check', 'label': 'Check'},
        {'id': 'cash', 'label': 'Cash'},
        {'id': 'credit_card', 'label': 'Credit Card'},
    ]

    products = [
        {'id': 'mezzuza', 'name': 'Mezzuza', 'price': '50'},
        {'id': 'lulav_etrog', 'name': 'Lulav & Etrog', 'price': '75'},
        {'id': 'memorial_candle', 'name': 'Memorial Candle', 'price': '5'},
        {'id': 'siddur', 'name': 'Siddur', 'price': '30'},
        {'id': 'machzor', 'name': 'Machzor', 'price': '40'},
        {'id': 'havdalah_set', 'name': 'Havdalah Set', 'price': '25'},
    ]

    kiddush_pricing = [
        {'id': 'standard', 'label': 'Standard Kiddush', 'price': '350', 'description': 'Stand-up kiddush with traditional spread'},
        {'id': 'sit-down', 'label': 'Sit-Down Kiddush', 'price': '500', 'description': 'Full sit-down kiddush meal'},
        {'id': 'deluxe', 'label': 'Deluxe Kiddush', 'price': '700', 'description': 'Premium deluxe kiddush experience'},
    ]

    seuda_pricing = [
        {'id': 'regular', 'label': 'Regular Seuda Shelishit', 'price': '250', 'description': 'Traditional Seuda Shelishit'},
        {'id': 'deluxe', 'label': 'Deluxe Seuda Shelishit', 'price': '400', 'description': 'Enhanced Seuda Shelishit spread'},
    ]

    settings = {
        'pledgeTypes': pledge_types,
        'occasions': occasions,
        'paymentMethods': payment_methods,
        'products': products,
        'kiddushPricing': kiddush_pricing,
        'seudaPricing': seuda_pricing,
    }

    for key, value in settings.items():
        settings_table.put_item(Item={
            'settingKey': key,
            'items': json.loads(json.dumps(value), parse_float=Decimal),
        })

    print(f"Uploaded {len(settings)} settings to stcd_settings")


# ========== MAIN ==========
if __name__ == '__main__':
    print("=== Uploading STCD data to DynamoDB ===\n")

    print("1. Uploading contacts...")
    name_to_id = upload_contacts()

    print("\n2. Uploading transactions & pledges...")
    upload_transactions(name_to_id)

    print("\n3. Uploading settings...")
    upload_settings()

    print("\n=== Done! ===")
